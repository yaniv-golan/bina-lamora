#!/usr/bin/env python3
"""Parse a Google Forms class-placement export (CSV/XLSX) into canonical students.json.

Usage:
  python parse_input.py export.csv --mapping mapping.json -o students.json
                        [--sheet "שם גיליון"]   # for multi-sheet xlsx

mapping.json may also include "class": "<header of a pre-assigned class column>" —
its raw values land on each student as "pre_class" (interpreted by parse_notes.py).
Gan names are normalized (trim, collapse spaces, drop leading 'גן '); merges warned.

mapping.json (see references/data-formats.md):
{
  "name": "<header of student name column>",
  "gan": "<header of kindergarten column>",
  "choices": ["<header choice 1>", "<header choice 2>", "<header choice 3>"],
  "gender": "<header or null>",
  "mishalevet": "<header or null>",
  "notes": "<header or null>",
  "mishalevet_names": ["optional list of student names flagged as mishalevet"]
}
Exit codes: 0 ok, 1 usage/file error, 2 mapping error.
"""
import argparse, csv, io, json, re, sys, unicodedata

TRUTHY = {"כן", "yes", "true", "1", "v", "+", "x", "משלבת", "כן."}

FINALS = str.maketrans("ךםןףץ", "כמנפצ")


def norm(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = "".join(ch for ch in s if not ("֑" <= ch <= "ׇ"))  # strip nikud/teamim
    s = s.translate(FINALS)
    s = re.sub(r"[\"'`’׳״]", "", s)
    s = re.sub(r"[^\w\s֐-׿-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def read_rows(path, sheet=None, expected_headers=None):
    """Returns (rows, warnings). For multi-sheet xlsx without --sheet, auto-picks the
    sheet whose header row matches the most expected headers; ambiguity is a JSON error
    listing sheet names (so the assistant can ask the user by name, not crash)."""
    warnings = []
    if path.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet is not None:
            if sheet not in wb.sheetnames:
                print(json.dumps({"error": f"sheet '{sheet}' not found",
                                  "sheets": wb.sheetnames}, ensure_ascii=False), file=sys.stderr)
                sys.exit(2)
            ws = wb[sheet]
        elif len(wb.sheetnames) == 1:
            ws = wb.active
        else:
            exp = {h.strip() for h in (expected_headers or []) if h}
            scores = {}
            for name in wb.sheetnames:
                head = next(wb[name].iter_rows(min_row=1, max_row=1, values_only=True), ()) or ()
                scores[name] = len(exp & {str(c).strip() for c in head if c is not None})
            best = max(scores.values(), default=0)
            winners = [nm for nm, sc in scores.items() if sc == best]
            if best == 0 or len(winners) > 1:
                print(json.dumps({"error": "multiple sheets — pass --sheet with one of these names",
                                  "sheets": wb.sheetnames, "header_matches": scores},
                                 ensure_ascii=False), file=sys.stderr)
                sys.exit(2)
            ws = wb[winners[0]]
            warnings.append(f"multi-sheet file: auto-selected sheet '{winners[0]}' "
                            f"({best} matching headers) — confirm with the user")
        rows = [["" if c is None else str(c).strip() for c in r] for r in ws.iter_rows(values_only=True)]
        return rows, warnings
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "cp1255", "iso-8859-8"):
        try:
            text = raw.decode(enc)
            # sanity: Hebrew files decoded with wrong codec show replacement junk
            if enc != "utf-8-sig" and "�" in text:
                continue
            break
        except UnicodeDecodeError:
            continue
    else:
        print(json.dumps({"error": "could not decode file"}), file=sys.stderr)
        sys.exit(1)
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t") if text.strip() else csv.excel
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text), dialect)], warnings


def norm_gan(raw):
    """Normalize a gan name: trim, collapse spaces, drop a leading 'גן ' prefix."""
    g = re.sub(r"\s+", " ", (raw or "").strip())
    g = re.sub(r"^גן\s+", "", g)
    return g


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input")
    ap.add_argument("--mapping", required=True)
    ap.add_argument("--sheet", help="sheet name for multi-sheet xlsx files")
    ap.add_argument("-o", "--output", default="students.json")
    a = ap.parse_args()

    mapping = json.load(open(a.mapping, encoding="utf-8"))
    expected = [mapping.get(k) for k in ("name", "gan", "gender", "mishalevet", "notes", "class")]
    expected += mapping.get("choices") or []
    rows, sheet_warnings = read_rows(a.input, sheet=a.sheet,
                                     expected_headers=[h for h in expected if h])
    if not rows:
        print(json.dumps({"error": "empty file"}), file=sys.stderr); sys.exit(1)
    header = rows[0]
    hidx = {h.strip(): i for i, h in enumerate(header)}
    # duplicate header text silently collapses in hidx (last wins) — if a MAPPED header
    # is duplicated, all its uses would read the same column and corrupt the data. Loud error.
    from collections import Counter as _Counter
    dup = {h for h, ct in _Counter(h.strip() for h in header if h and h.strip()).items() if ct > 1}
    mapped = {v for v in ([mapping.get(x) for x in ("name", "gan", "gender", "mishalevet", "notes", "class")]
                          + (mapping.get("choices") or [])) if v}
    bad_dup = sorted(dup & mapped)
    if bad_dup:
        print(json.dumps({"error": "duplicate column headers used in the mapping — rename the "
                          "columns in the file so each mapped header is unique",
                          "duplicates": bad_dup}, ensure_ascii=False), file=sys.stderr)
        sys.exit(2)

    def col(key, required=False):
        h = mapping.get(key)
        if h is None:
            if required:
                print(json.dumps({"error": f"mapping missing required key '{key}'"}), file=sys.stderr); sys.exit(2)
            return None
        if h not in hidx:
            print(json.dumps({"error": f"header not found: '{h}'", "available": header}), file=sys.stderr); sys.exit(2)
        return hidx[h]

    ci_name, ci_gan = col("name", True), col("gan", True)
    choice_headers = mapping.get("choices") or []
    if not choice_headers:
        print(json.dumps({"error": "mapping missing 'choices'"}), file=sys.stderr); sys.exit(2)
    ci_choices = []
    for h in choice_headers:
        if h not in hidx:
            print(json.dumps({"error": f"header not found: '{h}'", "available": header}), file=sys.stderr); sys.exit(2)
        ci_choices.append(hidx[h])
    ci_gender, ci_mish, ci_notes, ci_class = col("gender"), col("mishalevet"), col("notes"), col("class")

    students, warnings, seen = [], list(sheet_warnings), {}
    gan_merges = {}
    for rn, row in enumerate(rows[1:], start=2):
        if not any(c for c in row):
            continue
        def get(i):
            return row[i].strip() if i is not None and i < len(row) else ""
        name = re.sub(r"\s+", " ", get(ci_name)).strip()
        if not name:
            warnings.append(f"row {rn}: empty student name, skipped")
            continue
        key = norm(name)
        if key in seen:
            seen[key] += 1
            warnings.append(f"duplicate student name '{name}' — suffixed as '{name} ({seen[key]})'; ask the user to distinguish them (e.g., by gan)")
            name = f"{name} ({seen[key]})"
        else:
            seen[key] = 1
        raw_gan = get(ci_gan)
        gan = norm_gan(raw_gan) or "לא ידוע"
        if raw_gan and gan != raw_gan:
            gan_merges.setdefault(raw_gan, gan)
        s = {
            "id": f"s{len(students):03d}",
            "name": name,
            "gan": gan,
            "gender": get(ci_gender) or None,
            "mishalevet": norm(get(ci_mish)) in {norm(t) for t in TRUTHY} if ci_mish is not None else False,
            "raw_choices": [get(i) for i in ci_choices],
            "notes": get(ci_notes) or "",
            "pre_class": get(ci_class) or "",
        }
        students.append(s)
    for raw_g, g in sorted(gan_merges.items()):
        warnings.append(f"gan name normalized: '{raw_g}' → '{g}'")

    for mn in mapping.get("mishalevet_names", []) or []:
        matches = [s for s in students if norm(s["name"]) == norm(mn) or norm(s["name"]).startswith(norm(mn))]
        if len(matches) == 1:
            matches[0]["mishalevet"] = True
        elif not matches:
            warnings.append(f"mishalevet name '{mn}' not found in roster")
        else:
            warnings.append(f"mishalevet name '{mn}' matches multiple students: {[m['name'] for m in matches]} — resolve manually")

    out = {"students": students, "warnings": warnings,
           "gans": sorted({s["gan"] for s in students}),
           "counts": {"students": len(students), "mishalevet": sum(s["mishalevet"] for s in students)}}
    json.dump(out, open(a.output, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(json.dumps({"ok": True, "students": len(students), "gans": out["gans"],
                      "mishalevet": out["counts"]["mishalevet"], "warnings": warnings}, ensure_ascii=False))


if __name__ == "__main__":
    main()
