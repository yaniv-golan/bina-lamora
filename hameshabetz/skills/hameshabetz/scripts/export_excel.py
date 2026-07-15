#!/usr/bin/env python3
"""Export final class rosters to an RTL Hebrew Excel workbook.

Usage:
  python export_excel.py reconciled.json solution.json --config config.json -o shibutz.xlsx

Sheets: one per class + סיכום (summary) + לתשומת לב (issues).
Requires openpyxl.
"""
import argparse, json
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="DDD5C4")
BOLD = Font(bold=True)


def style_header(ws, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=1, column=i)
        c.font = BOLD
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="right")


def autowidth(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("reconciled")
    ap.add_argument("solution")
    ap.add_argument("--config", required=True)
    ap.add_argument("-o", "--output", default="shibutz.xlsx")
    a = ap.parse_args()

    data = json.load(open(a.reconciled, encoding="utf-8"))
    sol = json.load(open(a.solution, encoding="utf-8"))
    config = json.load(open(a.config, encoding="utf-8"))
    students = data["students"]
    by_id = {s["id"]: s for s in students}
    assign = {sid: int(c) for sid, c in sol["assignment"].items()}
    k = config["num_classes"]
    names = sol.get("class_names") or config.get("class_names") or [f"כיתה {i+1}" for i in range(k)]

    def sat_info(s):
        best, bestname = None, ""
        for ch in s.get("choices", []):
            if assign.get(s["id"]) == assign.get(ch["target"]):
                if best is None or ch["rank"] < best:
                    best, bestname = ch["rank"], by_id[ch["target"]]["name"]
        return best, bestname

    wb = Workbook()
    wb.remove(wb.active)
    for ci in range(k):
        ws = wb.create_sheet(title=names[ci][:31])
        ws.sheet_view.rightToLeft = True
        ws.append(["#", "שם", "גן", "מין", "משלבת", "חבר בכיתה (בחירה)", "הערות"])
        style_header(ws, 7)
        members = sorted((s for s in students if assign.get(s["id"]) == ci),
                         key=lambda s: s["name"])
        for i, s in enumerate(members, 1):
            best, bestname = sat_info(s)
            friend = f"{bestname} ({best})" if best else ("—" if s.get("choices") else "אין בחירות תקפות")
            ws.append([i, s["name"], s["gan"], s.get("gender") or "", "כן" if s.get("mishalevet") else "",
                       friend, s.get("notes", "")])
            if s.get("choices") and not best:
                ws.cell(row=ws.max_row, column=6).font = Font(color="CC0000", bold=True)
        autowidth(ws)

    # summary
    ws = wb.create_sheet(title="סיכום")
    ws.sheet_view.rightToLeft = True
    sizes = defaultdict(int); mish = defaultdict(int)
    gan_matrix = defaultdict(lambda: defaultdict(int))
    counts = {1: 0, 2: 0, 3: 0, "none": 0, "no_choices": 0}
    for s in students:
        c = assign.get(s["id"]); sizes[c] += 1
        if s.get("mishalevet"):
            mish[c] += 1
        gan_matrix[s["gan"]][c] += 1
        best, _ = sat_info(s)
        if best:
            counts[best] += 1
        elif s.get("choices"):
            counts["none"] += 1
        else:
            counts["no_choices"] += 1
    ws.append(["כיתה", "תלמידים", "משלבת"] + [f"גן {g}" for g in sorted(gan_matrix)])
    style_header(ws, 3 + len(gan_matrix))
    for ci in range(k):
        ws.append([names[ci], sizes[ci], mish[ci]] + [gan_matrix[g][ci] for g in sorted(gan_matrix)])
    ws.append([])
    ws.append(["קיבלו בחירה ראשונה", counts[1]])
    ws.append(["קיבלו בחירה שנייה", counts[2]])
    ws.append(["קיבלו בחירה שלישית", counts[3]])
    ws.append(["ללא אף חבר שבחרו", counts["none"]])
    ws.append(["ללא בחירות תקפות בטופס", counts["no_choices"]])
    autowidth(ws)

    # issues
    ws = wb.create_sheet(title="לתשומת לב")
    ws.sheet_view.rightToLeft = True
    ws.append(["סוג", "פרטים"])
    style_header(ws, 2)
    rec = data.get("reconciliation", {})
    for s in students:
        best, _ = sat_info(s)
        if s.get("choices") and not best:
            ws.append(["ילד ללא אף חבר שבחר בכיתתו", f"{s['name']} ({names[assign[s['id']]]})"])
    for kd in rec.get("zero_matched_choices", []):
        ws.append(["אף בחירה לא זוהתה במאגר", f"{kd['name']} (גן {kd['gan']})"])
    for kd in rec.get("no_choices_on_form", []):
        ws.append(["לא מילא בחירות בטופס", f"{kd['name']} (גן {kd['gan']})"])
    for u in rec.get("unmatched", []):
        ws.append(["שם שלא זוהה", f"{u['chooser_name']} כתב/ה: '{u['raw']}'"])
    for w in data.get("warnings", []):
        ws.append(["אזהרת קליטה", w])
    autowidth(ws)

    wb.save(a.output)
    print(json.dumps({"ok": True, "output": a.output, "sheets": [s.title for s in wb.worksheets]},
                     ensure_ascii=False))


if __name__ == "__main__":
    main()
